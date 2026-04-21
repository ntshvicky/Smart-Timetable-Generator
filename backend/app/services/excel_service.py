from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models import (
    AcademicYear,
    ClassSubjectRequirement,
    PeriodDefinition,
    SchoolClass,
    SchoolWorkingDay,
    Section,
    Subject,
    Teacher,
    TeacherAvailability,
    TeacherSubjectClassMap,
    Timetable,
    TimetableEntry,
    UploadBatch,
    UploadErrorLog,
)

DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]


@dataclass
class ImportErrorItem:
    sheet_name: str
    row_number: int
    column_name: str
    message: str


class ExcelService:
    sheet_aliases = {
        "Class Subject Requirement": ["Class Subject Weekly Requirement"],
    }
    required_sheets = {
        "Classes": ["class_name", "section_name", "class_display_name"],
        "Subjects": ["subject_code", "subject_name", "category"],
        "Teachers": ["teacher_code", "teacher_name", "max_periods_per_day", "max_consecutive_periods"],
        "Teacher-Class-Subject Mapping": ["teacher_code", "class_name", "section_name", "subject_code"],
        "Teacher Availability": ["teacher_code", "day", "period_number", "available_yes_no"],
        "Class Subject Requirement": [
            "class_name",
            "section_name",
            "subject_code",
            "periods_per_week",
            "preferred_first_half",
            "preferred_last_period",
            "avoid_consecutive_yes_no",
        ],
        "School Settings": ["academic_year", "working_days", "periods_per_day", "lunch_after_period"],
    }

    def __init__(self, db: Session):
        self.db = db

    def build_template(self) -> BytesIO:
        wb = Workbook()
        wb.remove(wb.active)
        examples: dict[str, list[list[Any]]] = {
            "Classes": [["5", "A", "Class 5A"], ["5", "B", "Class 5B"], ["6", "A", "Class 6A"]],
            "Subjects": [["MATH", "Math", "heavy"], ["ENG", "English", "language"], ["SCI", "Science", "heavy"], ["PT", "PT", "activity"]],
            "Teachers": [
                ["T-RAVI", "Ravi", 8, 3], ["T-MAYA", "Maya", 8, 3], ["T-OM", "Om", 8, 3],
                ["T-RINA", "Rina", 8, 3], ["T-NEEL", "Neel", 8, 3], ["T-ILA", "Ila", 8, 3],
                ["T-ANU", "Anu", 8, 2], ["T-SARA", "Sara", 8, 2], ["T-KAI", "Kai", 8, 2],
            ],
            "Teacher-Class-Subject Mapping": [
                ["T-RAVI", "5", "A", "MATH"], ["T-MAYA", "5", "A", "MATH"], ["T-OM", "5", "A", "MATH"], ["T-RINA", "5", "A", "ENG"], ["T-NEEL", "5", "A", "ENG"], ["T-ILA", "5", "A", "ENG"], ["T-ANU", "5", "A", "SCI"], ["T-SARA", "5", "A", "SCI"], ["T-KAI", "5", "A", "SCI"],
                ["T-RAVI", "5", "B", "MATH"], ["T-MAYA", "5", "B", "MATH"], ["T-OM", "5", "B", "MATH"], ["T-RINA", "5", "B", "ENG"], ["T-NEEL", "5", "B", "ENG"], ["T-ILA", "5", "B", "ENG"], ["T-ANU", "5", "B", "SCI"], ["T-SARA", "5", "B", "SCI"], ["T-KAI", "5", "B", "SCI"],
                ["T-RAVI", "6", "A", "MATH"], ["T-MAYA", "6", "A", "MATH"], ["T-OM", "6", "A", "MATH"], ["T-RINA", "6", "A", "ENG"], ["T-NEEL", "6", "A", "ENG"], ["T-ILA", "6", "A", "ENG"], ["T-ANU", "6", "A", "SCI"], ["T-SARA", "6", "A", "SCI"], ["T-KAI", "6", "A", "SCI"],
            ],
            "Teacher Availability": [["T-RAVI", "Monday", 1, "yes"], ["T-RAVI", "Wednesday", 4, "no"]],
            "Class Subject Requirement": [
                ["5", "A", "MATH", 5, "yes", "no", "yes"], ["5", "A", "ENG", 5, "no", "no", "yes"], ["5", "A", "SCI", 4, "no", "no", "no"],
                ["5", "B", "MATH", 5, "yes", "no", "yes"], ["5", "B", "ENG", 5, "no", "no", "yes"], ["5", "B", "SCI", 4, "no", "no", "no"],
                ["6", "A", "MATH", 5, "yes", "no", "yes"], ["6", "A", "ENG", 5, "no", "no", "yes"], ["6", "A", "SCI", 4, "no", "no", "no"],
            ],
            "School Settings": [["2026-2027", "Monday,Tuesday,Wednesday,Thursday,Friday", 8, 4]],
        }
        for sheet, headers in self.required_sheets.items():
            ws = wb.create_sheet(sheet)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2563EB")
            for row in examples.get(sheet, []):
                ws.append(row)
            for col_idx, header in enumerate(headers, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = max(16, len(header) + 2)
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    def import_workbook(self, school_id: int, filename: str, content: bytes) -> tuple[UploadBatch, list[ImportErrorItem]]:
        batch = UploadBatch(school_id=school_id, filename=filename, status="validating")
        self.db.add(batch)
        self.db.flush()
        wb = load_workbook(BytesIO(content), data_only=True)
        errors = self._validate_workbook(wb)
        if errors:
            batch.status = "failed"
            for item in errors:
                self.db.add(UploadErrorLog(upload_batch_id=batch.id, **item.__dict__))
            self.db.commit()
            return batch, errors

        self._replace_school_data(school_id)
        self._import_settings(school_id, self._rows(wb, "School Settings"))
        self._import_classes(school_id, self._rows(wb, "Classes"))
        self._import_subjects(school_id, self._rows(wb, "Subjects"))
        self._import_teachers(school_id, self._rows(wb, "Teachers"))
        self.db.flush()
        self._import_mappings(school_id, self._rows(wb, "Teacher-Class-Subject Mapping"))
        self._import_availability(school_id, self._rows(wb, "Teacher Availability"))
        self._import_requirements(school_id, self._rows(wb, "Class Subject Requirement"))
        batch.status = "imported"
        self.db.commit()
        return batch, []

    def _validate_workbook(self, wb) -> list[ImportErrorItem]:
        errors: list[ImportErrorItem] = []
        for sheet, headers in self.required_sheets.items():
            actual_sheet = self._sheet_name(wb, sheet)
            if not actual_sheet:
                errors.append(ImportErrorItem(sheet, 0, "", "Missing required sheet"))
                continue
            actual = [str(cell.value).strip() if cell.value is not None else "" for cell in wb[actual_sheet][1]]
            for header in headers:
                if header not in actual:
                    errors.append(ImportErrorItem(sheet, 1, header, "Missing required column"))
        if errors:
            return errors

        row_sets = {sheet: self._rows(wb, sheet) for sheet in self.required_sheets}
        teachers = self._unique_values(row_sets["Teachers"], "teacher_code", errors, "Teachers")
        subjects = self._unique_values(row_sets["Subjects"], "subject_code", errors, "Subjects")
        sections = {(str(r.get("class_name")).strip(), str(r.get("section_name")).strip()) for _, r in row_sets["Classes"]}
        for row_num, row in row_sets["Teacher-Class-Subject Mapping"]:
            self._ref(row, row_num, "Teacher-Class-Subject Mapping", "teacher_code", teachers, errors)
            self._ref(row, row_num, "Teacher-Class-Subject Mapping", "subject_code", subjects, errors)
            if (str(row.get("class_name")).strip(), str(row.get("section_name")).strip()) not in sections:
                errors.append(ImportErrorItem("Teacher-Class-Subject Mapping", row_num, "class_name", "Unknown class/section"))
        for row_num, row in row_sets["Class Subject Requirement"]:
            self._ref(row, row_num, "Class Subject Requirement", "subject_code", subjects, errors)
            if int(row.get("periods_per_week") or 0) <= 0:
                errors.append(ImportErrorItem("Class Subject Requirement", row_num, "periods_per_week", "Must be greater than zero"))
        return errors

    def _rows(self, wb, sheet_name: str) -> list[tuple[int, dict[str, Any]]]:
        actual_sheet = self._sheet_name(wb, sheet_name)
        if not actual_sheet:
            return []
        ws = wb[actual_sheet]
        headers = [str(cell.value).strip() for cell in ws[1]]
        rows = []
        for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if all(value is None for value in values):
                continue
            rows.append((idx, dict(zip(headers, values))))
        return rows

    def _sheet_name(self, wb, sheet_name: str) -> str | None:
        if sheet_name in wb.sheetnames:
            return sheet_name
        for alias in self.sheet_aliases.get(sheet_name, []):
            if alias in wb.sheetnames:
                return alias
        return None

    def _unique_values(self, rows: list[tuple[int, dict[str, Any]]], col: str, errors: list[ImportErrorItem], sheet: str) -> set[str]:
        seen: set[str] = set()
        for row_num, row in rows:
            value = str(row.get(col) or "").strip()
            if not value:
                errors.append(ImportErrorItem(sheet, row_num, col, "Required value missing"))
            elif value in seen:
                errors.append(ImportErrorItem(sheet, row_num, col, "Duplicate value"))
            seen.add(value)
        return seen

    def _ref(self, row: dict[str, Any], row_num: int, sheet: str, col: str, valid: set[str], errors: list[ImportErrorItem]) -> None:
        if str(row.get(col) or "").strip() not in valid:
            errors.append(ImportErrorItem(sheet, row_num, col, f"Unknown {col}"))

    def _replace_school_data(self, school_id: int) -> None:
        for model in [TimetableEntry, Timetable, TeacherAvailability, TeacherSubjectClassMap, ClassSubjectRequirement, PeriodDefinition, SchoolWorkingDay, AcademicYear, Section, SchoolClass, Subject, Teacher]:
            self.db.execute(delete(model).where(model.school_id == school_id))

    def _import_settings(self, school_id: int, rows: list[tuple[int, dict[str, Any]]]) -> None:
        row = rows[0][1] if rows else {}
        academic_year = str(row.get("academic_year") or "2026-2027")
        working_days = [d.strip() for d in str(row.get("working_days") or "Monday,Tuesday,Wednesday,Thursday,Friday").split(",") if d.strip()]
        periods = int(row.get("periods_per_day") or 8)
        lunch_after = int(row.get("lunch_after_period") or 0)
        self.db.add(AcademicYear(school_id=school_id, name=academic_year, is_active=True))
        for idx, day in enumerate(working_days, 1):
            self.db.add(SchoolWorkingDay(school_id=school_id, day=day, sort_order=idx))
        for period in range(1, periods + 1):
            self.db.add(PeriodDefinition(school_id=school_id, period_number=period, label=f"Period {period}", is_break=(period == lunch_after + 1 if lunch_after else False)))

    def _import_classes(self, school_id: int, rows: list[tuple[int, dict[str, Any]]]) -> None:
        class_cache: dict[str, SchoolClass] = {}
        for _, row in rows:
            class_name = str(row["class_name"]).strip()
            school_class = class_cache.get(class_name)
            if not school_class:
                school_class = SchoolClass(school_id=school_id, name=class_name, display_name=f"Class {class_name}")
                self.db.add(school_class)
                self.db.flush()
                class_cache[class_name] = school_class
            section = str(row["section_name"]).strip()
            self.db.add(Section(school_id=school_id, class_id=school_class.id, name=section, display_name=str(row.get("class_display_name") or f"{class_name}{section}")))

    def _import_subjects(self, school_id: int, rows: list[tuple[int, dict[str, Any]]]) -> None:
        for _, row in rows:
            self.db.add(Subject(school_id=school_id, code=str(row["subject_code"]).strip(), name=str(row["subject_name"]).strip(), category=str(row.get("category") or "general").strip()))

    def _import_teachers(self, school_id: int, rows: list[tuple[int, dict[str, Any]]]) -> None:
        for _, row in rows:
            self.db.add(Teacher(school_id=school_id, code=str(row["teacher_code"]).strip(), name=str(row["teacher_name"]).strip(), max_periods_per_day=int(row.get("max_periods_per_day") or 6), max_consecutive_periods=int(row.get("max_consecutive_periods") or 3)))

    def _lookup(self, model, school_id: int, attr: str, value: str):
        return self.db.scalar(select(model).where(model.school_id == school_id, getattr(model, attr) == value))

    def _section(self, school_id: int, class_name: str, section_name: str) -> Section:
        school_class = self._lookup(SchoolClass, school_id, "name", class_name)
        return self.db.scalar(select(Section).where(Section.school_id == school_id, Section.class_id == school_class.id, Section.name == section_name))

    def _import_mappings(self, school_id: int, rows: list[tuple[int, dict[str, Any]]]) -> None:
        for _, row in rows:
            self.db.add(TeacherSubjectClassMap(
                school_id=school_id,
                teacher_id=self._lookup(Teacher, school_id, "code", str(row["teacher_code"]).strip()).id,
                subject_id=self._lookup(Subject, school_id, "code", str(row["subject_code"]).strip()).id,
                section_id=self._section(school_id, str(row["class_name"]).strip(), str(row["section_name"]).strip()).id,
            ))

    def _import_availability(self, school_id: int, rows: list[tuple[int, dict[str, Any]]]) -> None:
        for _, row in rows:
            teacher = self._lookup(Teacher, school_id, "code", str(row["teacher_code"]).strip())
            self.db.add(TeacherAvailability(school_id=school_id, teacher_id=teacher.id, day=str(row["day"]).strip(), period_number=int(row["period_number"]), is_available=str(row["available_yes_no"]).strip().lower() in {"yes", "y", "true", "1"}))

    def _import_requirements(self, school_id: int, rows: list[tuple[int, dict[str, Any]]]) -> None:
        for _, row in rows:
            self.db.add(ClassSubjectRequirement(
                school_id=school_id,
                section_id=self._section(school_id, str(row["class_name"]).strip(), str(row["section_name"]).strip()).id,
                subject_id=self._lookup(Subject, school_id, "code", str(row["subject_code"]).strip()).id,
                periods_per_week=int(row["periods_per_week"]),
                preferred_first_half=str(row.get("preferred_first_half") or "").lower() in {"yes", "y", "true", "1"},
                preferred_last_period=str(row.get("preferred_last_period") or "").lower() in {"yes", "y", "true", "1"},
                avoid_consecutive=str(row.get("avoid_consecutive_yes_no") or "").lower() in {"yes", "y", "true", "1"},
            ))
